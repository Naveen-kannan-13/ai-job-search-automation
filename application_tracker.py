"""
Application tracker module for recording job applications.
"""
from datetime import datetime
from pathlib import Path


class ApplicationTracker:
    """Track job applications and save to Excel."""
    
    def __init__(self, file_path="job_applications.xlsx"):
        """
        Initialize the application tracker.
        
        Args:
            file_path (str): Path to Excel file for storing applications
        """
        self.file_path = file_path
        self.applications = []
        self._load_existing()
    
    def _load_existing(self):
        """Load existing applications from Excel file."""
        if Path(self.file_path).exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.file_path)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Skip empty rows
                        self.applications.append({
                            'company': row[0],
                            'position': row[1],
                            'link': row[2],
                            'match_score': row[3],
                            'status': row[4],
                            'date_applied': row[5],
                            'notes': row[6]
                        })
                wb.close()
            except ImportError:
                print("openpyxl not installed. Install with: pip install openpyxl")
    
    def add_application(self, company, position, link, match_score, status="Applied", notes=""):
        """
        Add a new job application to the tracker.
        
        Args:
            company (str): Company name
            position (str): Job position/title
            link (str): Job posting link
            match_score (int): Match score from job matcher
            status (str): Application status (Applied, Interview, Rejected, etc.)
            notes (str): Additional notes about the application
        """
        application = {
            'company': company,
            'position': position,
            'link': link,
            'match_score': match_score,
            'status': status,
            'date_applied': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'notes': notes
        }
        self.applications.append(application)
        return application
    
    def save_to_excel(self):
        """Save all applications to Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create or load workbook
            try:
                wb = openpyxl.load_workbook(self.file_path)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)  # Clear existing data
            except:
                wb = openpyxl.Workbook()
                ws = wb.active
            
            # Add headers
            headers = ['Company', 'Position', 'Link', 'Match Score (%)', 'Status', 
                      'Date Applied', 'Notes']
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data rows
            for app in self.applications:
                ws.append([
                    app['company'],
                    app['position'],
                    app['link'],
                    app['match_score'],
                    app['status'],
                    app['date_applied'],
                    app['notes']
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 30
            
            # Add color coding for match scores
            for row in ws.iter_rows(min_row=2, max_row=len(self.applications) + 1, min_col=4, max_col=4):
                cell = row[0]
                if cell.value:
                    score = int(cell.value)
                    if score >= 80:
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    elif score >= 60:
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            wb.save(self.file_path)
            wb.close()
            print(f"✓ Saved {len(self.applications)} applications to {self.file_path}")
            return True
        
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return False
    
    def get_stats(self):
        """Get application statistics."""
        total = len(self.applications)
        by_status = {}
        
        for app in self.applications:
            status = app['status']
            by_status[status] = by_status.get(status, 0) + 1
        
        avg_score = sum(app['match_score'] for app in self.applications) / total if total > 0 else 0
        
        return {
            'total_applications': total,
            'by_status': by_status,
            'average_match_score': round(avg_score, 1)
        }
    
    def print_summary(self):
        """Print a summary of applications."""
        stats = self.get_stats()
        print("\n" + "="*50)
        print("APPLICATION SUMMARY")
        print("="*50)
        print(f"Total Applications: {stats['total_applications']}")
        print(f"Average Match Score: {stats['average_match_score']}%")
        print("\nBy Status:")
        for status, count in stats['by_status'].items():
            print(f"  - {status}: {count}")
        print("="*50 + "\n")


if __name__ == "__main__":
    # Test the tracker
    tracker = ApplicationTracker()
    
    tracker.add_application(
        company="Stripe",
        position="Senior Frontend Engineer",
        link="https://jobs.lever.co/stripe/job-id",
        match_score=85,
        notes="Great match for skills"
    )
    
    tracker.save_to_excel()
    tracker.print_summary()
